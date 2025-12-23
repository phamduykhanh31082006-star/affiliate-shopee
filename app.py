from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import requests


app = Flask(__name__)

EXCEL_FILE = "data.xlsx"

def save_to_excel(data):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Thời gian",
            "Họ tên",
            "Số điện thoại",
            "Email",
            "Link Shopee / MXH"
        ])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    wb.save(EXCEL_FILE)

@app.route("/", methods=["GET"])
def home():
    return render_template("index.html")

import requests

@app.route("/register", methods=["POST"])
def register():
    name = request.form.get("name")
    phone = request.form.get("phone")
    email = request.form.get("email")
    link = request.form.get("link")

    google_script_url = "https://script.google.com/macros/s/AKfycbxLiLZFJGiPobVRyrJOOC5W9LdHJCoTWIjCiIA-hHe7YemHeMhOFGKk75lDRQP4gacw/exec"

    payload = {
        "name": name,
        "phone": phone,
        "email": email,
        "link": link
    }

    try:
        requests.post(google_script_url, json=payload, timeout=10)
        return render_template("success.html", name=name)
    except Exception as e:
        return "Có lỗi xảy ra, vui lòng thử lại"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)

